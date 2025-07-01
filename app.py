# Colab-friendly FBA & Shipping Label Extraction + Template Fill with Auto-Detect & Barcode Decoding
# ---------------------------------------------------------------
# Before running this cell, install dependencies in a separate Colab cell:
# !apt-get update -qq && apt-get install -qq poppler-utils libzbar0 libsm6 libxext6
# !pip install pdf2image pyzbar pytesseract opencv-python-headless openpyxl

import os, glob, re
import pandas as pd
import pytesseract
import cv2
import numpy as np
from pdf2image import convert_from_path
from pyzbar.pyzbar import decode as zbar_decode
from google.colab import files

# 1) Auto-detect PDFs in current directory
def detect_pdfs():
    pdfs = glob.glob("*.pdf")
    fba = [f for f in pdfs if f.upper().startswith('FBA')]
    ship = [f for f in pdfs if f.upper().startswith('LABELS')]
    if not (fba and ship):
        raise FileNotFoundError("Need one 'FBA*.pdf' and one 'Labels*.pdf' in cwd.")
    return fba[0], ship[0]

fba_pdf, ship_pdf = detect_pdfs()
template_xl = 'Amazon_Tracking_Uploads_Template.xlsx'
base = os.path.splitext(os.path.basename(fba_pdf))[0]
shipment_id = base.split('-')[0]
output_xl = f"{shipment_id}_tracking_upload.xlsx"

print(f"Using FBA PDF: {fba_pdf}\nUsing Shipping PDF: {ship_pdf}\nTemplate: {template_xl}\n")

# 2) OCR configurations
OCR_FULL_CFG  = "--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
OCR_LINE_CFG  = "--psm 7 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ- "

# 3) Barcode decode helper
def decode_barcode(img):
    decoded = zbar_decode(img)
    if decoded:
        data = decoded[0].data.decode('utf-8')
        code = re.sub(r'[^0-9A-Za-z]', '', data)
        return ' '.join(code[i:i+4] for i in range(0, len(code), 4))
    return None

# 4) Extraction function
def extract_box_and_tracking(pdf_path, prefix):
    pages = convert_from_path(pdf_path, dpi=300)
    records = []
    for page_num, pil_img in enumerate(pages, start=1):
        img_file = f"{prefix}_page_{page_num}.png"
        pil_img.save(img_file)
        img = cv2.imread(img_file)
        h, w = img.shape[:2]

        # FBA box-ID region
        y1, y2 = int(h*0.2), int(h*0.45)
        band = img[y1:y2, int(w*0.05):int(w*0.95)]
        box_id = decode_barcode(band)
        if not box_id:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            txt = pytesseract.image_to_string(gray, config=OCR_FULL_CFG).upper()
            m = re.search(r'FBA[0-9A-Z]+', txt)
            box_id = m.group(0) if m else ""

        # UPS tracking region
        y3, y4 = int(h*0.5), int(h*0.8)
        band2 = img[y3:y4, int(w*0.05):int(w*0.95)]
        tracking = decode_barcode(band2)
        if not tracking:
            gray2 = cv2.cvtColor(band2, cv2.COLOR_BGR2GRAY)
            _, bw = cv2.threshold(gray2, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
            txt2 = pytesseract.image_to_string(bw, config=OCR_LINE_CFG).upper()
            m2 = re.search(r'1Z[0-9A-Z]{8,}', txt2)
            if m2:
                raw = re.sub(r'[^0-9A-Z]', '', m2.group(0))
                tracking = ' '.join(raw[i:i+4] for i in range(0, len(raw), 4))
            else:
                tracking = ""
                print(f"[DEBUG] Page {page_num} tracking OCR: '{txt2.strip()}'")

        records.append({'page': page_num, 'box_id': box_id, 'tracking': tracking})
    return pd.DataFrame(records)

# 5) Run extraction and merge
boxes_df = extract_box_and_tracking(fba_pdf, 'fba')[['page','box_id']]
track_df = extract_box_and_tracking(ship_pdf, 'ship')[['page','tracking']]
merged = pd.merge(boxes_df, track_df, on='page')
print(merged)

# 6) Fill Excel template and save
from openpyxl import load_workbook
wb = load_workbook(template_xl)
ws = wb.active

# Shipment ID in B4
ws.cell(row=4, column=2, value=shipment_id)

# Write cleaned values starting at row 8
start_row = 8
for idx, row in merged.iterrows():
    r = start_row + idx
    clean_box   = row['box_id'].replace(' ', '')
    clean_track = row['tracking'].replace(' ', '')
    ws.cell(row=r, column=1, value=clean_box)
    ws.cell(row=r, column=2, value=clean_track)

wb.save(output_xl)
print(f"✅ Filled template saved as: {output_xl}")

# 7) Download the result file
files.download(output_xl)
